#!/usr/bin/env python3
"""
Excel Report Generator for Reloj Laboral
Generates Excel reports with charts for daily, weekly, and monthly data.
"""

import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("Installing required packages...")
    os.system("pip install pandas openpyxl")
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows


def style_header(ws):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_daily_report(data: list, output_path: str):
    """Generate daily Excel report"""
    if not data:
        print("No data to generate report")
        return
    
    df = pd.DataFrame(data)
    
    if df.empty:
        print("DataFrame is empty")
        return
    
    timestamp_col = 'timestamp'
    if timestamp_col in df.columns:
        df['date'] = pd.to_datetime(df[timestamp_col]).dt.date
        df['time'] = pd.to_datetime(df[timestamp_col]).dt.time
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Diario"
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    style_header(ws)
    auto_adjust_columns(ws)
    
    wb.save(output_path)
    print(f"Daily report saved: {output_path}")


def create_weekly_report(data: dict, output_path: str):
    """Generate weekly Excel report with summary"""
    if not data or 'dailySummary' not in data:
        print("No data to generate report")
        return
    
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen Semanal"
    
    headers = ["Fecha", "Horas Trabajadas"]
    ws_summary.append(headers)
    
    daily_summary = data.get('dailySummary', {})
    total_hours = 0
    
    for date, hours in sorted(daily_summary.items()):
        ws_summary.append([date, round(hours, 2)])
        total_hours += hours
    
    ws_summary.append(["Total", round(total_hours, 2)])
    
    style_header(ws_summary)
    
    last_row = ws_summary.max_row
    ws_summary.append([""])
    ws_summary.append([f"Semana: {data.get('startDate', '')} - {data.get('endDate', '')}"])
    
    if len(daily_summary) > 0:
        ws_chart = wb.create_sheet("Gráfico")
        chart = BarChart()
        chart.type = "col"
        chart.title = "Horas por Día"
        
        data_ref = Reference(ws_summary, min_col=2, min_row=1, max_row=len(daily_summary))
        cats_ref = Reference(ws_summary, min_col=1, min_row=2, max_row=len(daily_summary))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_chart.add_chart(chart, "E5")
    
    wb.save(output_path)
    print(f"Weekly report saved: {output_path}")


def create_monthly_report(data: dict, output_path: str):
    """Generate monthly Excel report"""
    if not data or 'dailySummary' not in data:
        print("No data to generate report")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Mensual"
    
    headers = ["Fecha", "Horas Trabajadas"]
    ws.append(headers)
    
    daily_summary = data.get('dailySummary', {})
    total_hours = 0
    
    for date, hours in sorted(daily_summary.items()):
        ws.append([date, round(hours, 2)])
        total_hours += hours
    
    ws.append([""])
    ws.append(["Total Horas", round(total_hours, 2)])
    
    style_header(ws)
    auto_adjust_columns(ws)
    
    wb.save(output_path)
    print(f"Monthly report saved: {output_path}")


def main():
    """Main function for CLI usage"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate Excel reports for Reloj Laboral")
    parser.add_argument('--type', choices=['daily', 'weekly', 'monthly'], required=True)
    parser.add_argument('--data', required=True, help='JSON data or path to JSON file')
    parser.add_argument('--output', required=True, help='Output file path')
    
    args = parser.parse_args()
    
    if os.path.isfile(args.data):
        with open(args.data, 'r') as f:
            import json
            data = json.load(f)
    else:
        import json
        data = json.loads(args.data)
    
    if args.type == 'daily':
        create_daily_report(data.get('clockings', []), args.output)
    elif args.type == 'weekly':
        create_weekly_report(data, args.output)
    elif args.type == 'monthly':
        create_monthly_report(data, args.output)


if __name__ == '__main__':
    main()